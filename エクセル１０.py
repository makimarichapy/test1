# -*- coding: utf-8 -*-
"""
このモジュールは Excel ファイルを解析し、
表(テーブル)および表以外のテキストを自動抽出する機能。

主な構成要素:
- TableRegion: 検出された表領域を表すデータクラス
- ExcelDetector: Excel ファイル全体を解析し FileResult を生成するメインクラス
"""

import os
import openpyxl
from typing import List, Dict, Tuple, Optional, Union
from dataclasses import dataclass, field


# ==========================
# 定数定義(設定セクション)
# ==========================
EXCLUDE_HEADERS = ['aaa', 'bbb', 'ccc']  # 除外対象ヘッダのリスト
MIN_TABLE_ROWS = 2  # 表と認識する最小行数
MIN_TABLE_COLS = 2  # 表と認識する最小列数

# 読み込み制限(性能確保のため)
MAX_READ_ROWS = 500  # 最大500行まで処理
MAX_READ_COLS = 500  # 最大500列まで処理

# ヘッダ文字長の制限(長過ぎる場合は省略記号に置換)
MAX_HEADER_LENGTH = 20


#########################
# 検出された表の情報を保持するデータクラス
#########################
@dataclass
class TableRegion:
    start_row: int   # 表の開始行番号
    start_col: int   # 表の開始列番号
    end_row: int     # 表の終了行番号
    end_col: int     # 表の終了列番号
    header_type: str   # ヘッダの種類(yoko_hedder、tate_hedder、yoko_multi_hedder、tate_multi_hedder)
    header_rows: List[int]  # ヘッダ行(横型ヘッダ)
    header_cols: List[int]  # ヘッダ列(縦型ヘッダ)
    confidence: float       # 検出信頼度(0.0～1.0)


#########################
# Excelセルの書式情報を保持するデータクラス
#########################
@dataclass
class CellFormat:
    is_bold: bool        # 太字かどうか
    has_bg_color: bool   # 背景色が設定されているか
    has_border: bool     # 罫線が設定されているか


#########################
# ワークシートのキャッシュ構造
#########################
@dataclass
class WorksheetCache:
    """セルの値と書式をキャッシュして再利用可能にする"""
    max_row: int  # 読み込み上限を反映した最大行
    max_col: int  # 読み込み上限を反映した最大列
    cell_formats: Dict[Tuple[int, int], CellFormat]  # (row, col)ごとの書式情報
    cell_values: Dict[Tuple[int, int], any]          # (row, col)ごとの値
    cell_values_str: Dict[Tuple[int, int], str]      # (row, col)ごとの文字列化された値(最適化用)
    merged_cells: set  # 結合セルの座標集合
    merged_cell_parents: Dict[Tuple[int, int], Tuple[int, int]]  # 結合セルの親座標マップ(最適化用)


#########################
# FileResult構造体群
#########################
@dataclass
class DataElement:
    """表またはテキストデータの単位情報"""
    container_file: str      # 内包ファイル名(通常は空文字)
    sheet_name: str          # Excelシート名
    table_number: int        # 表番号(1開始)またはテキスト番号
    data_type: str           # "table"または"text"
    content: Union[List[List[str]], str]  # 表は2次元リスト、テキストは文字列


@dataclass
class FileResult:
    """ファイル単位の解析結果をまとめる"""
    context: Dict[str, str] = field(default_factory=dict)  # メタ情報(例: {"root_file": "sample.xlsx"})
    elements: List[DataElement] = field(default_factory=list)  # 含まれるDataElementのリスト


#########################
# Excelシート解析クラス
#########################
class ExcelDetector:
    """
    Excelファイル全体を解析し、シート単位で表・テキストを検出する。
    書式情報(太字・背景色)を利用し、表構造を推定する。
    """

    def __init__(self):
        """設定値を初期化"""
        self.min_table_rows = MIN_TABLE_ROWS   # 表とみなす最小行数
        self.min_table_cols = MIN_TABLE_COLS   # 表とみなす最小列数
        self.max_read_rows = MAX_READ_ROWS     # 最大読み込み行数
        self.max_read_cols = MAX_READ_COLS     # 最大読み込み列数
        self.exclude_headers = EXCLUDE_HEADERS # 除外ヘッダリスト
        self.max_header_length = MAX_HEADER_LENGTH # ヘッダ文字長制限

    # ============================================================
    # メイン処理:Excelファイル全体を処理
    # ============================================================
    def read_excel2(self, file_path: str) -> FileResult:
        """
        指定されたExcelファイルを開き、全シートを解析する。
        表領域および表外テキストを抽出してFileResult構造で返す。
        """
        # 出力結果を格納するFileResultを初期化
        file_result = FileResult(
            context={"root_file": os.path.basename(file_path)},
            elements=[]
        )
        
        try:
            # Excelブックを読み込み(値と書式を分けてロード)
            workbook_values = openpyxl.load_workbook(file_path, data_only=True)   # 計算結果のみ
            workbook_formats = openpyxl.load_workbook(file_path, data_only=False) # 書式付き
        except Exception as e:
            # ファイルオープン失敗時はエラーログ出力して空の結果を返す
            print(f"エラー: ファイル '{file_path}' の読み込みに失敗: {e}")
            return file_result

        # 全シートを順に解析
        for sheet_name in workbook_values.sheetnames:

            worksheet_values = workbook_values[sheet_name]   # 値のみシート
            worksheet_formats = workbook_formats[sheet_name] # 書式付きシート
            
            # キャッシュを1回だけ生成して使い回す
            cache = self._extract_worksheet_cache(worksheet_values, worksheet_formats)
            
            # ① 表検出
            tables = self.detect_excel_tables(cache)

            # ② 表以外のテキストを抽出
            text_elements = self._extract_text_elements(cache, tables)
            
            # テキスト領域が存在する場合はDataElementとして追加
            if text_elements:
                element = DataElement(
                    container_file="",       #原則、空白
                    sheet_name=sheet_name,   #シート名を設定
                    table_number=0,          #テキストは0番扱い
                    data_type="text",        #タイプ:テキスト
                    content=text_elements[0] #テキスト文書を設定
                )
                file_result.elements.append(element)
            
            # 検出された各表を処理
            for table_number, table in enumerate(tables, start=1):
                try:
                    # 表データを抽出
                    table_data = self.get_table_data(cache, table)
                    
                    # 空データはスキップ
                    if not table_data or len(table_data) == 0:
                        continue
                    
                    # 表情報をDataElementとして追加
                    element = DataElement(
                        container_file="",         #原則、空白
                        sheet_name=sheet_name,     #シート名を設定
                        table_number=table_number, #連番
                        data_type="table",         #タイプ:table
                        content=table_data         #テーブルデータを設定
                    )
                    file_result.elements.append(element)
                    
                except Exception as e:
                    # 個別の表抽出失敗は警告のみで継続
                    print(f"警告: シート '{sheet_name}' の表 {table_number} のデータ取得に失敗: {e}")
                    continue
        
        return file_result


    def _extract_text_elements(self, cache: WorksheetCache, tables: List[TableRegion]) -> List[str]:
        """
        表以外に存在する単独テキスト(タイトル・注釈など)を抽出する。
        """
        text_parts = []  # 抽出テキストの一時格納リスト
        
        # 検出済み表の座標をセット化(テキストとの重複を防ぐ)
        table_cells = set()
        for table in tables:
            for row in range(table.start_row, table.end_row + 1):
                for col in range(table.start_col, table.end_col + 1):
                    table_cells.add((row, col))
        
        # シート全セルをスキャンして表外のテキストを収集
        for row_idx in range(1, cache.max_row + 1):
            for col_idx in range(1, cache.max_col + 1):
                if (row_idx, col_idx) in table_cells:
                    continue  # 表領域内セルはスキップ
                
                # ★最適化ポイント7: 文字列化済みの値を使用★
                text_content = cache.cell_values_str.get((row_idx, col_idx), "")
                # 非空セルをテキストとして抽出
                if text_content:
                    text_parts.append(text_content)
        
        # 複数のテキストを改行結合して1つの文字列にまとめる
        if text_parts:
            combined_text = '\n'.join(text_parts)
            return [combined_text]
        
        return []


    # ============================================================
    # 準メイン処理:テーブルを検出
    # ============================================================
    def detect_excel_tables(self, cache: WorksheetCache) -> List[TableRegion]:
        """
        シート内に存在する表領域を検出する。
        書式やデータ密度をもとに、横型・縦型・混在を判定。
        """

        # シートの構造方向を推定(横主体・縦主体など)
        sheet_orientation = self._determine_sheet_orientation(cache)

        # 構造方向に応じて優先検出モードを切り替え
        if sheet_orientation == "横型主体":
            high_confidence_tables = self._detect_yoko_tables_primary(cache)
        elif sheet_orientation == "縦型主体":
            high_confidence_tables = self._detect_tate_tables_primary(cache)
        else:
            # 不明または混在時は厳格ルールで検出
            high_confidence_tables = self._detect_with_strict_rules(cache)

        # 高信頼度検出結果がある場合は重複統合して返す
        if high_confidence_tables:
            return self._merge_overlapping_tables(high_confidence_tables)

        # 次に中程度ルールで検出
        medium_confidence_tables = self._detect_with_medium_rules(cache)
        if medium_confidence_tables:
            return self._merge_overlapping_tables(medium_confidence_tables)

        # どの検出にも該当しない場合は全体を表とみなす(フォールバック)
        return self._fallback_whole_sheet_as_table(cache)


    def _determine_sheet_orientation(self, cache: WorksheetCache) -> str:
        """
        シート全体の特徴から「横型」、「縦型」、「混在」を推定する。

        判定基準:
        - 書式の多い行数 vs 書式の多い列数
        - データ行数と列数の比率(aspect ratio)
        """
        # ★最適化ポイント2: 書式チェックとデータチェックを1回のループで実行★
        formatted_rows = 0
        formatted_cols = 0
        data_rows = 0
        data_cols_set = set()
        
        # 先頭20行の書式をチェック
        for row in range(1, min(cache.max_row + 1, 21)):
            if self._check_row_formatting(cache, row):
                formatted_rows += 1
            
            # データの有無もチェック
            has_data = False
            for col in range(1, cache.max_col + 1):
                if cache.cell_values_str.get((row, col)):
                    has_data = True
                    data_cols_set.add(col)
            if has_data:
                data_rows += 1
        
        # 先頭20列の書式をチェック
        for col in range(1, min(cache.max_col + 1, 21)):
            if self._check_col_formatting(cache, col):
                formatted_cols += 1
        
        data_cols = len(data_cols_set)
        
        # データ比率を計算
        aspect_ratio = data_rows / max(data_cols, 1)
        
        # スコア算出(行書式数+形状補正)
        yoko_score = formatted_rows + (1 if aspect_ratio < 0.5 else 0)
        tate_score = formatted_cols + (1 if aspect_ratio > 2.0 else 0)
        
        # 最終判定(しきい値比1.5倍)
        if yoko_score > tate_score * 1.5:
            return "横型主体"
        elif tate_score > yoko_score * 1.5:
            return "縦型主体"
        else:
            return "混在または不明"


    def _detect_yoko_tables_primary(self, cache: WorksheetCache) -> List[TableRegion]:
        """
        横型(行方向)に展開された表を優先的に検出する。
        太字行や複数行ヘッダをもとに領域を確定する。
        """
        tables: List[TableRegion] = []  # 検出結果格納リスト
        processed_rows = set()          # 処理済み行を記録(重複防止)

        # まず複数行ヘッダ形式を検出
        multi_row_tables = self._detect_multi_row_headers(cache)
        for table in multi_row_tables:
            for row in range(table.start_row, table.end_row + 1):
                processed_rows.add(row)
        tables.extend(multi_row_tables)

        # 残りの行を順次確認してヘッダ候補を探索
        for row_idx in range(1, cache.max_row + 1):
            if row_idx in processed_rows:
                continue  # 既に含まれる行はスキップ
                
            if not self._is_strong_yoko_header(cache, row_idx):
                continue  # 強いヘッダ判定を満たさない場合スキップ

            # ヘッダ候補行から表領域を抽出
            table_region = self._extract_yoko_table(cache, row_idx)
            if table_region and self._meets_minimum_size(table_region):
                table_region.confidence = 0.9  # 高信頼度とみなす
                
                # 検出行を登録して再検出防止
                for row in range(table_region.start_row, table_region.end_row + 1):
                    processed_rows.add(row)
                    
                tables.append(table_region)

        return tables


    def _detect_tate_tables_primary(self, cache: WorksheetCache) -> List[TableRegion]:
        """
        縦型(列方向)に展開された表を優先的に検出する。
        書式付き列・データ右連続性などを利用。
        """
        tables: List[TableRegion] = []
        processed_cols = set()  # 処理済み列を記録

        for col_idx in range(1, cache.max_col + 1):
            if col_idx in processed_cols:
                continue
                
            if not self._is_strong_tate_header(cache, col_idx):
                continue  # 強い縦ヘッダ条件を満たさない列はスキップ

            # 縦型表領域を抽出
            table_region = self._extract_tate_table(cache, col_idx)
            if table_region and self._meets_minimum_size(table_region):
                table_region.confidence = 0.9  # 高信頼度設定
                
                for col in range(table_region.start_col, table_region.end_col + 1):
                    processed_cols.add(col)

                # 縦→横に転置して統一形式に変換
                table_region = self._transpose_tate_table(table_region)
                tables.append(table_region)

        return tables


    def get_table_data(self, cache: WorksheetCache, table_region: TableRegion) -> List[List[str]]:
        """
        検出された表領域からセルデータを2次元リストとして抽出する。

        - ヘッダ文字長制限や除外ヘッダの適用を行う
        - 複数行ヘッダは '__' で結合する
        """
        
        # 縦型表の場合は転置データ抽出ルートへ
        if table_region.header_type == "yoko_hedder" and table_region.header_cols:
            return self._get_transposed_tate_table_data(cache, table_region)
        
        # 複数行ヘッダの場合は結合処理を実施
        if table_region.header_type == "yoko_multi_hedder":
            merged_header = self._merge_multi_row_headers(cache, table_region)
        else:
            # 通常ヘッダ行の取得
            merged_header = []
            for col in range(table_region.start_col, table_region.end_col + 1):
                # ★最適化ポイント7: 文字列化済みの値を使用★
                header_value = cache.cell_values_str.get((table_region.start_row, col), "")
                merged_header.append(header_value)
        
        # ヘッダ文字数制限を適用
        merged_header = self._apply_header_length_limit(merged_header)
        # 除外リストに該当する列インデックスを取得
        exclude_indices = self._get_exclude_column_indices(merged_header)
        # 除外列を削除したヘッダを生成
        filtered_header = [h for i, h in enumerate(merged_header) if i not in exclude_indices]
        
        # ヘッダが空の場合はスキップ
        if not filtered_header or all(not h.strip() for h in filtered_header):
            return []
        
        data = [filtered_header]  # 先頭にヘッダ行を追加
        
        # データ部分を1行ずつ収集
        for row in range(table_region.start_row + len(table_region.header_rows), table_region.end_row + 1):
            row_data = []
            for col_idx, col in enumerate(range(table_region.start_col, table_region.end_col + 1)):
                if col_idx not in exclude_indices:
                    # ★最適化ポイント7: 文字列化済みの値を使用★
                    cell_str = cache.cell_values_str.get((row, col), "")
                    row_data.append(cell_str)
            data.append(row_data)
        
        return data



    def _get_transposed_tate_table_data(self, cache: WorksheetCache, table_region: TableRegion) -> List[List[str]]:
        """転置された縦型表のデータを正しい形式で取得する"""
        original_start_row = table_region.start_col
        original_end_row = table_region.end_col
        original_start_col = table_region.start_row
        original_end_col = table_region.end_row
        
        header = []
        for row in range(original_start_row, original_end_row + 1):
            # ★最適化ポイント7: 文字列化済みの値を使用★
            header_value = cache.cell_values_str.get((row, original_start_col), "")
            header.append(header_value)
        
        header = self._apply_header_length_limit(header)
        data = [header]
        
        for col in range(original_start_col + 1, original_end_col + 1):
            row_data = []
            for row in range(original_start_row, original_end_row + 1):
                # ★最適化ポイント7: 文字列化済みの値を使用★
                cell_str = cache.cell_values_str.get((row, col), "")
                row_data.append(cell_str)
            data.append(row_data)
        
        return data


    def _extract_worksheet_cache(self, worksheet_values, worksheet_formats) -> WorksheetCache:
        """
        ワークシートから書式情報と値情報を一括抽出し、キャッシュに格納する。
        """
    
        # ===== 1. 読み取り範囲の決定 =====
        max_row = min(worksheet_values.max_row, self.max_read_rows)
        max_col = min(worksheet_values.max_column, self.max_read_cols)
    
        # ===== 2. データ格納用の辞書を初期化 =====
        # キー: (row, col) のタプル
        cell_formats = {}       # 書式情報(太字、背景色、罫線など)
        cell_values_str = {}    # 文字列化されたセル値
        merged_cells_set = set()       # 結合セルの全座標
        merged_cell_parents = {}       # 結合セルの親座標マップ
    
        # ===== 3. 結合セル情報の抽出 =====
        for merged_range in worksheet_formats.merged_cells:
            parent_row = merged_range.min_row
            parent_col = merged_range.min_col
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cells_set.add((row, col))
                    merged_cell_parents[(row, col)] = (parent_row, parent_col)
    
        # ===== 4. セルの値（文字列）を一括取得 =====
        for row in worksheet_values.iter_rows(min_row=1, max_row=max_row,
                                              min_col=1, max_col=max_col,
                                              values_only=False):  # セルオブジェクト取得
            for cell in row:
                coord = (cell.row, cell.column)
                # 値を文字列化（Noneは空文字に変換）
                value = str(cell.value).strip() if cell.value is not None else ""
                cell_values_str[coord] = value
    
        # ===== 5. セルの書式情報を一括取得 =====
        for row in worksheet_formats.iter_rows(min_row=1, max_row=max_row,
                                               min_col=1, max_col=max_col,
                                               values_only=False):
            for cell in row:
                coord = (cell.row, cell.column)
                # 文字列値が空でないセルのみ処理（メモリ最適化）
                if cell_values_str.get(coord):
                    cell_formats[coord] = CellFormat(
                        is_bold=bool(getattr(cell.font, "bold", False)),
                        has_bg_color=self._has_background_color(cell),
                        has_border=self._has_border(cell)
                    )
    
        # ===== 6. WorksheetCache にまとめて返す =====
        return WorksheetCache(
            max_row=max_row,
            max_col=max_col,
            cell_formats=cell_formats,
            cell_values_str=cell_values_str,
            merged_cells=merged_cells_set,
            merged_cell_parents=merged_cell_parents
        )




    def _detect_with_strict_rules(self, cache: WorksheetCache) -> List[TableRegion]:
        """
        厳格な条件に基づいて表(TableRegion)を検出する。
    
        概要:
            1. 複数行ヘッダを持つ表（multi-row header）を検出
            2. その後、横ヘッダを抽出
            3. 次に、縦ヘッダを抽出
            4. 重複する表領域はスキップし、信頼度(confidence)を付与
    
        戻り値:
            List[TableRegion]: 検出された表領域のリスト
        """
    
        # === 検出結果格納用 ===
        tables: List[TableRegion] = []   # 検出した表領域の一覧
        processed_rows = set()           # 既に表として処理済みの行インデックス集合
    
        # === 1. 複数行ヘッダを持つ表を優先検出 ===
        # 例: ヘッダが2行以上続くタイプの表
        multi_row_tables = self._detect_multi_row_headers(cache)
        
        # 検出された複数行ヘッダ表の行を「処理済み」に登録
        for table in multi_row_tables:
            for row in range(table.start_row, table.end_row + 1):
                processed_rows.add(row)
    
        # 表リストに追加
        tables.extend(multi_row_tables)
    
        # === 2. 横方向ヘッダ（行方向の見出し）に基づく表を検出 ===
        # 上から下へ1行ずつチェック
        for row_idx in range(1, cache.max_row + 1):
    
            # 既に別の表に含まれる行はスキップ
            if row_idx in processed_rows:
                continue
    
            # 横方向のヘッダ構造が強い行かを判定（例: 太字・背景色・セル値の分布）
            if not self._is_strong_yoko_header(cache, row_idx):
                continue
    
            # 横型表の抽出処理（行単位の構造解析）
            table_region = self._extract_yoko_table(cache, row_idx)
    
            # 表として成立しており、かつ最小サイズ要件を満たす場合のみ採用
            if table_region and self._meets_minimum_size(table_region):
                # 厳格検出なので高い信頼度(0.9)を設定
                table_region.confidence = 0.9
    
                # この表に含まれる行を「処理済み」として登録
                for row in range(table_region.start_row, table_region.end_row + 1):
                    processed_rows.add(row)
    
                # 検出結果に追加
                tables.append(table_region)
    
        # === 3. 縦方向ヘッダ（列方向の見出し）に基づく表を検出 ===
        for col_idx in range(1, cache.max_col + 1):
    
            # 強い縦ヘッダ（列見出し）を持つ列かどうかを判定
            if not self._is_strong_tate_header(cache, col_idx):
                continue
    
            # 縦型表を抽出（列単位の構造解析）
            table_region = self._extract_tate_table(cache, col_idx)
    
            # 表として成立し、最小サイズを満たす場合のみ採用
            if table_region and self._meets_minimum_size(table_region):
    
                # 既に他の表と行が重複していないかを確認
                overlaps = False
                for row in range(table_region.start_row, table_region.end_row + 1):
                    if row in processed_rows:
                        overlaps = True
                        break
    
                # 他の表と重複していない場合のみ登録
                if not overlaps:
                    table_region.confidence = 0.9
    
                    # この表の行を「処理済み」にマーク
                    for row in range(table_region.start_row, table_region.end_row + 1):
                        processed_rows.add(row)
    
                    # 縦方向の表は後処理として転置（横型に変換）
                    table_region = self._transpose_tate_table(table_region)
    
                    # 検出結果に追加
                    tables.append(table_region)
    
        # === 4. すべての検出結果を返す ===
        return tables




    def _detect_multi_row_headers(self, cache: WorksheetCache) -> List[TableRegion]:
        """
        連続する書式付き行を「複数行ヘッダ」として検出する。
    
        概要:
            - セル書式（太字・背景色・罫線など）が連続している行群を検出し、
              それらを複数行ヘッダとして1つの表領域(TableRegion)にまとめる。
            - ヘッダが2行以上連続している場合のみ表候補とみなす。
            - ヘッダの直下にデータが連続して存在することも確認し、
              実際に表が展開されていると判断できる場合に採用する。
    
        戻り値:
            List[TableRegion]: 検出された複数行ヘッダ構造の表領域リスト
        """
    
        tables = []  # 検出結果格納リスト
    
        # === 1. 行ごとに先頭行候補を探索 ===
        #   - 各行を「ヘッダ開始行」と仮定してチェックを行う
        for start_row in range(1, cache.max_row):
    
            header_rows = []  # 書式付き行が連続している行番号のリスト
    
            # === 2. 最大3行分のヘッダ候補を確認 ===
            #   - 3行連続までを上限とし、それ以上は別表として扱う
            for offset in range(6):
                check_row = start_row + offset
    
                # ワークシートの最終行を超えた場合は終了
                if check_row > cache.max_row:
                    break
    
                # 対象行が書式付き（ヘッダっぽい）か判定
                if self._check_row_formatting(cache, check_row):
                    header_rows.append(check_row)
                else:
                    # 書式付き行の連続が途切れたら中断
                    break
    
            # === 3. 書式付き行が2行以上連続している場合のみ採用 ===
            if len(header_rows) >= 2:
    
                # 直下にデータ行が連続して存在するかを確認
                # （ヘッダだけで終わる装飾領域を除外するため）
                if self._check_data_continuity_below(cache, header_rows[-1]):
    
                    # === 4. 表の左右端を推定 ===
                    #   - 左端: 最初に値が出現する列
                    #   - 右端: 連続して値が存在する最終列
                    start_col = self._find_table_start_col(cache, header_rows[0])
                    end_col = self._find_table_end_col(cache, header_rows[0])
    
                    # === 5. 表の下端を推定 ===
                    #   - データが途切れるまで探索して最終行を特定
                    end_row = self._find_table_end_row(
                        cache,
                        header_rows[-1],
                        start_col,
                        end_col
                    )
    
                    # === 6. 表領域オブジェクトを生成 ===
                    table_region = TableRegion(
                        start_row=header_rows[0],      # ヘッダ開始行
                        start_col=start_col,           # 表の左端列
                        end_row=end_row,               # 表の下端行
                        end_col=end_col,               # 表の右端列
                        header_type="yoko_multi_hedder",  # ヘッダ種別（横方向・複数行）
                        header_rows=header_rows,       # 検出されたヘッダ行リスト
                        header_cols=[],                # 縦方向ヘッダはなし
                        confidence=0.9                 # 検出信頼度（厳格検出なので高め）
                    )
    
                    # === 7. 最小サイズ要件を満たす場合のみ採用 ===
                    if self._meets_minimum_size(table_region):
                        tables.append(table_region)
    
        # === 8. 全検出結果を返す ===
        return tables


    def _is_strong_yoko_header(self, cache: WorksheetCache, row_idx: int) -> bool:
        """指定行が横型ヘッダとして適格か、厳格な条件で判定する"""
        row_values = self._get_row_values(cache, row_idx)
        if len(row_values) < self.min_table_cols:
            return False
        
        has_formatting = self._check_row_formatting(cache, row_idx)
        # ★最適化ポイント3: 重複チェックを改善★
        is_unique = self._check_header_uniqueness(row_values)
        has_data_below = self._check_data_continuity_below(cache, row_idx)
        
        return has_formatting and is_unique and has_data_below

    def _is_strong_tate_header(self, cache: WorksheetCache, col_idx: int) -> bool:
        """指定列が縦型ヘッダとして適格か、厳格な条件で判定する"""
        col_values = self._get_col_values(cache, col_idx)
        non_empty_values = [v for v in col_values if v]
        
        if len(non_empty_values) < self.min_table_rows:
            return False
        
        has_formatting = self._check_col_formatting(cache, col_idx)
        # ★最適化ポイント3: 重複チェックを改善★
        is_unique = self._check_header_uniqueness(non_empty_values)
        has_data_right = self._check_data_continuity_right(cache, col_idx)
        
        return has_formatting and is_unique and has_data_right

    def _detect_with_medium_rules(self, cache: WorksheetCache) -> List[TableRegion]:
        """中程度のルールでの表検出"""
        tables = []

        for row_idx in range(1, cache.max_row + 1):
            if self._is_medium_yoko_header(cache, row_idx):
                table_region = self._extract_yoko_table(cache, row_idx)

                if table_region and self._meets_minimum_size(table_region):
                    table_region.confidence = 0.7
                    tables.append(table_region)

        return tables

    def _is_medium_yoko_header(self, cache: WorksheetCache, row_idx: int) -> bool:
        """指定行が横型ヘッダとして適格か、中程度の条件で判定する"""
        row_values = self._get_row_values(cache, row_idx)
        if len(row_values) < self.min_table_cols:
            return False
        
        has_formatting = self._check_row_formatting(cache, row_idx)
        # ★最適化ポイント3: 重複チェックを改善★
        is_unique = self._check_header_uniqueness(row_values)
        has_data_below = self._check_data_continuity_below(cache, row_idx)
        
        pattern_a = has_formatting and has_data_below
        pattern_b = is_unique and has_data_below
        
        return pattern_a or pattern_b

    def _fallback_whole_sheet_as_table(self, cache: WorksheetCache) -> List[TableRegion]:
        """シート全体を1つの表として扱うフォールバック処理"""
        has_data = False
        for value_str in cache.cell_values_str.values():
            if value_str:
                has_data = True
                break
        
        if not has_data:
            return []
        
        if cache.max_row < self.min_table_rows or cache.max_col < self.min_table_cols:
            return []

        return [TableRegion(
            start_row=1,
            start_col=1,
            end_row=cache.max_row,
            end_col=cache.max_col,
            header_type="yoko_hedder",
            header_rows=[1],
            header_cols=[],
            confidence=0.3
        )]

    def _extract_yoko_table(self, cache: WorksheetCache, header_row: int) -> Optional[TableRegion]:
        """横型表の境界を抽出"""
        start_col = self._find_table_start_col(cache, header_row)
        end_col = self._find_table_end_col(cache, header_row)
        end_row = self._find_table_end_row(cache, header_row, start_col, end_col)
        
        if end_row - header_row < 2:
            return None
        
        return TableRegion(
            start_row=header_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
            header_type="yoko_hedder",
            header_rows=[header_row],
            header_cols=[],
            confidence=0.0
        )

    def _extract_tate_table(self, cache: WorksheetCache, header_col: int) -> Optional[TableRegion]:
        """縦型表の境界を抽出"""
        start_row = self._find_table_start_row_for_tate(cache, header_col)
        end_row = self._find_table_end_row_for_tate(cache, header_col)
        end_col = self._find_table_end_col_for_tate(cache, header_col, start_row, end_row)
    
        total_cols = end_col - header_col + 1
        if total_cols < self.min_table_cols:
            return None
    
        return TableRegion(
            start_row=start_row,
            start_col=header_col,
            end_row=end_row,
            end_col=end_col,
            header_type="tate_hedder",
            header_rows=[],
            header_cols=[header_col],
            confidence=0.0
        )

    def _transpose_tate_table(self, table_region: TableRegion) -> TableRegion:
        """縦型表を横型表に転置する"""
        return TableRegion(
            start_row=table_region.start_col,
            start_col=table_region.start_row,
            end_row=table_region.end_col,
            end_col=table_region.end_row,
            header_type="yoko_hedder",
            header_rows=[table_region.start_col],
            header_cols=table_region.header_cols,
            confidence=table_region.confidence
        )

    def _find_table_start_col(self, cache: WorksheetCache, row_idx: int) -> int:
        """表の開始列を検出"""
        for col_idx in range(1, cache.max_col + 1):
            if cache.cell_values_str.get((row_idx, col_idx)):
                return col_idx
        return 1

    def _find_table_end_col(self, cache: WorksheetCache, row_idx: int) -> int:
        """表の終了列を検出"""
        last_col = 1
        empty_count = 0
        
        for col_idx in range(1, cache.max_col + 1):
            if cache.cell_values_str.get((row_idx, col_idx)):
                last_col = col_idx
                empty_count = 0
            else:
                empty_count += 1
                if empty_count >= 3:
                    break
        
        return last_col

    def _find_table_end_row(self, cache: WorksheetCache, start_row: int, 
                           start_col: int, end_col: int) -> int:
        """表の終了行を検出"""
        last_row = start_row
        empty_count = 0
        
        for row_idx in range(start_row + 1, cache.max_row + 1):
            has_data = False
            for col_idx in range(start_col, end_col + 1):
                if cache.cell_values_str.get((row_idx, col_idx)):
                    has_data = True
                    break
            
            if has_data:
                last_row = row_idx
                empty_count = 0
            else:
                empty_count += 1
                if empty_count >= 2:
                    break
        
        return last_row

    def _find_table_start_row_for_tate(self, cache: WorksheetCache, col_idx: int) -> int:
        """縦型表の開始行を検出"""
        for row_idx in range(1, cache.max_row + 1):
            if cache.cell_values_str.get((row_idx, col_idx)):
                return row_idx
        return 1

    def _find_table_end_row_for_tate(self, cache: WorksheetCache, col_idx: int) -> int:
        """縦型表の終了行を検出"""
        last_row = 1
        for row_idx in range(1, cache.max_row + 1):
            if cache.cell_values_str.get((row_idx, col_idx)):
                last_row = row_idx
        return last_row

    def _find_table_end_col_for_tate(self, cache: WorksheetCache, start_col: int, 
                                     start_row: int, end_row: int) -> int:
        """縦型表の終了列を検出"""
        last_col = start_col
        empty_count = 0
        
        for col_idx in range(start_col + 1, cache.max_col + 1):
            has_data = False
            for row_idx in range(start_row, end_row + 1):
                if cache.cell_values_str.get((row_idx, col_idx)):
                    has_data = True
                    break
            
            if has_data:
                last_col = col_idx
                empty_count = 0
            else:
                empty_count += 1
                if empty_count >= 2:
                    break
        
        return last_col

    def _check_row_formatting(self, cache: WorksheetCache, row_idx: int) -> bool:
        """指定行の書式設定をチェック"""
        formatted_count = 0
        total_cells = 0
        
        for col_idx in range(1, cache.max_col + 1):
            if cache.cell_values.get((row_idx, col_idx)) is not None:
                total_cells += 1
                cell_format = cache.cell_formats.get((row_idx, col_idx))
                if cell_format and (cell_format.is_bold or cell_format.has_bg_color):
                    formatted_count += 1
        
        return total_cells > 0 and (formatted_count / total_cells) >= 0.5

    def _check_col_formatting(self, cache: WorksheetCache, col_idx: int) -> bool:
        """指定列の書式設定をチェック"""
        formatted_count = 0
        total_cells = 0
        
        for row_idx in range(1, cache.max_row + 1):
            if cache.cell_values.get((row_idx, col_idx)) is not None:
                total_cells += 1
                cell_format = cache.cell_formats.get((row_idx, col_idx))
                if cell_format and (cell_format.is_bold or cell_format.has_bg_color):
                    formatted_count += 1
        
        return total_cells > 0 and (formatted_count / total_cells) >= 0.5

    def _check_header_uniqueness(self, values: List) -> bool:
        """
        ヘッダ値の一意性をチェック
        ★最適化ポイント3: 文字列変換を1回のみに削減★
        """
        # 文字列化を1回だけ実行
        non_empty_str_values = []
        for v in values:
            if v:  # 既に文字列化済みの値を前提
                non_empty_str_values.append(v if isinstance(v, str) else str(v))
        
        if len(non_empty_str_values) < 2:
            return False
        
        unique_count = len(set(non_empty_str_values))
        total_count = len(non_empty_str_values)
        duplicate_rate = 1 - (unique_count / total_count)
        
        return duplicate_rate <= 0.2

    def _check_data_continuity_below(self, cache: WorksheetCache, row_idx: int) -> bool:
        """下方向のデータ連続性をチェック"""
        data_rows = 0
        
        for check_row in range(row_idx + 1, min(row_idx + 10, cache.max_row + 1)):
            row_values = self._get_row_values(cache, check_row)
            non_empty = [v for v in row_values if v]
            
            if len(non_empty) >= self.min_table_cols - 1:
                data_rows += 1
            else:
                break
        
        return data_rows >= 2

    def _check_data_continuity_right(self, cache: WorksheetCache, col_idx: int) -> bool:
        """右方向のデータ連続性をチェック"""
        data_cols = 0
        
        header_start_row = None
        header_end_row = None
        for row in range(1, cache.max_row + 1):
            if cache.cell_values_str.get((row, col_idx)):
                if header_start_row is None:
                    header_start_row = row
                header_end_row = row
        
        if header_start_row is None:
            return False
        
        for check_col in range(col_idx + 1, min(col_idx + 11, cache.max_col + 1)):
            has_data = False
            for row in range(header_start_row, header_end_row + 1):
                if cache.cell_values_str.get((row, check_col)):
                    has_data = True
                    break
            
            if has_data:
                data_cols += 1
            else:
                break
        
        return data_cols >= (self.min_table_cols - 1)

    def _merge_overlapping_tables(self, tables: List[TableRegion]) -> List[TableRegion]:
        """
        重複する表領域をマージ
        ★最適化ポイント6: 空間分割で効率化（ただし今回はシンプルな改善版）★
        """
        if not tables:
            return []
        
        # 信頼度と開始行でソート
        sorted_tables = sorted(tables, key=lambda t: (-t.confidence, t.start_row))
        
        result = []
        for table in sorted_tables:
            is_overlapping = False
            
            # 既存の表と重複チェック
            for existing in result:
                if self._tables_overlap(table, existing):
                    is_overlapping = True
                    break
            
            # 重複していなければ追加
            if not is_overlapping:
                result.append(table)
        
        # 開始行順にソート
        result.sort(key=lambda t: t.start_row)
        
        return result

    def _tables_overlap(self, table1: TableRegion, table2: TableRegion) -> bool:
        """2つの表領域が重複しているかチェック"""
        row_overlap = not (
            table1.end_row < table2.start_row or table2.end_row < table1.start_row
        )

        col_overlap = not (
            table1.end_col < table2.start_col or table2.end_col < table1.start_col
        )

        return row_overlap and col_overlap

    def _merge_multi_row_headers(self, cache: WorksheetCache, table_region: TableRegion) -> List[str]:
        """
        複数行ヘッダを1行に結合(結合セル考慮版)
        ★最適化ポイント4: 結合セル親座標を事前計算済みマップから取得★
        """
        merged_headers = []
        
        for col in range(table_region.start_col, table_region.end_col + 1):
            parts = []
            
            for row in table_region.header_rows:
                coord = (row, col)
                
                # 結合セルの場合は親セルの値を取得
                if coord in cache.merged_cells:
                    parent_coord = cache.merged_cell_parents.get(coord, coord)
                    part = cache.cell_values_str.get(parent_coord, "")
                else:
                    part = cache.cell_values_str.get(coord, "")
                
                parts.append(part)
            
            merged_header = '__'.join(parts)
            merged_headers.append(merged_header)
        
        return merged_headers

    def _get_exclude_column_indices(self, headers: List[str]) -> set:
        """除外対象のヘッダに合致する列インデックスを取得"""
        exclude_indices = set()
        
        normalized_excludes = [h.lower().strip() for h in self.exclude_headers]
        
        for idx, header in enumerate(headers):
            normalized_header = header.lower().strip()
            if normalized_header in normalized_excludes:
                exclude_indices.add(idx)
        
        return exclude_indices

    def _get_row_values(self, cache: WorksheetCache, row_idx: int) -> List[str]:
        """
        指定行の全セル値を取得
        ★最適化ポイント7: 文字列化済みの値を使用★
        """
        return [cache.cell_values_str.get((row_idx, col), "") 
                for col in range(1, cache.max_col + 1)]

    def _get_col_values(self, cache: WorksheetCache, col_idx: int) -> List[str]:
        """
        指定列の全セル値を取得
        ★最適化ポイント7: 文字列化済みの値を使用★
        """
        return [cache.cell_values_str.get((row, col_idx), "") 
                for row in range(1, cache.max_row + 1)]

    def _meets_minimum_size(self, table_region: TableRegion) -> bool:
        """表が最小サイズ要件を満たすかチェック"""
        rows = table_region.end_row - table_region.start_row + 1
        cols = table_region.end_col - table_region.start_col + 1
        return rows >= self.min_table_rows and cols >= self.min_table_cols

    def _has_background_color(self, cell) -> bool:
        """セルに背景色が設定されているかをチェック"""
        try:
            if not hasattr(cell, 'fill') or not hasattr(cell.fill, 'start_color'):
                return False
            
            rgb = getattr(cell.fill.start_color, 'rgb', None)
            fill_type = getattr(cell.fill, 'fill_type', None)
            
            return (rgb is not None and 
                    rgb not in ['00000000', 'FFFFFFFF'] and 
                    fill_type is not None and 
                    fill_type != 'none')
        except:
            return False

    def _has_border(self, cell) -> bool:
        """セルに罫線が設定されているかをチェック"""
        try:
            if not hasattr(cell, 'border'):
                return False
            
            border = cell.border
            return any([
                getattr(border.top, 'style', None),
                getattr(border.bottom, 'style', None),
                getattr(border.left, 'style', None),
                getattr(border.right, 'style', None)
            ])
        except:
            return False

    def _apply_header_length_limit(self, headers: List[str]) -> List[str]:
        """ヘッダ文字列長を制限する"""
        limited_headers = []
        for header in headers:
            if len(header) > self.max_header_length:
                limited_headers.append(header[:self.max_header_length] + "...")
            else:
                limited_headers.append(header)
        return limited_headers


if __name__ == "__main__":
    import sys
    
    file_path = "./test.xlsx"
    
    if not os.path.exists(file_path):
        print(f"エラー: テストファイル '{file_path}' が見つかりません")
        sys.exit(1)
    
    print("=" * 80)
    print("ExcelDetector 動作確認テスト")
    print("=" * 80)
    
    detector = ExcelDetector()
    
    print(f"\n📂 処理中: {file_path}")
    file_result = detector.read_excel2(file_path)
    
    print(f"\n{'='*80}")
    print(f"📘 対象ファイル : {file_result.context['root_file']}")
    print(f"📊 検出された要素数 : {len(file_result.elements)}")
    print(f"{'='*80}\n")
    
    for element in file_result.elements:
        print("-" * 80)
        
        if element.data_type == "text":
            print(f"[テキスト(表以外のテキスト)]")
            print(f"  📄 シート名     : {element.sheet_name}")
            print(f"  📝 データタイプ : {element.data_type}")
            print(f"  📃 内容(先頭200文字):")
            
            content_preview = element.content[:200]
            if len(element.content) > 200:
                content_preview += "..."
            
            for line in content_preview.split('\n'):
                if line.strip():
                    print(f"     {line}")
        
        elif element.data_type == "table":
            print(f"[表 {element.table_number}]")
            print(f"  📄 シート名     : {element.sheet_name}")
            print(f"  📝 データタイプ : {element.data_type}")
            print(f"  📏 行数         : {len(element.content)}(ヘッダ含む)")
            print(f"  📏 列数         : {len(element.content[0]) if element.content else 0}")
            
            if element.content:
                print(f"  📹 ヘッダ行:")
                header = element.content[0]
                for i, h in enumerate(header, 1):
                    display_header = h if len(h) <= 30 else h[:27] + "..."
                    print(f"     列{i}: {display_header}")
                
                if len(element.content) > 1:
                    print(f"  📹 データサンプル(最初の3行):")
                    for row_idx, row in enumerate(element.content[1:4], 1):
                        sample_cols = row[:3]
                        print(f"     行{row_idx}: {sample_cols}")
                        if len(row) > 3:
                            print(f"            ... 他 {len(row) - 3} 列")
                else:
                    print(f"  📹 データ行なし")
    
    print("\n" + "=" * 80)
    print("✅ ExcelDetector の動作確認が完了しました")
    print("=" * 80)