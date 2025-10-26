# -*- coding: utf-8 -*-
"""
Created on Sat Oct 25 15:43:25 2025

@author: maki
"""

"""
Excelファイルから表とテキストを自動抽出するプログラム
パスワードファイル検出システムの一機能として、表構造とテキストを識別・抽出
"""

import os

import openpyxl
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional, Any


# ==================== 設定値・定数 ====================
# 除外対象ヘッダ（部分一致）
EXCLUDE_HEADERS = ['ユーザー名', 'bbb', 'ccc']

# 除外対象シート名（部分一致）
EXCLUDE_SHEETS = ['sample', 'サンプル', 'ccc']

# 表として認識する最小サイズ
MIN_TABLE_ROWS = 2  # ヘッダ除く最小データ行数
MIN_TABLE_COLS = 2  # ヘッダ除く最小データ行数

# 最大読み込み範囲（性能対策）
MAX_READ_ROWS = 500
MAX_READ_COLS = 500

# ヘッダの最大文字長
MAX_HEADER_LENGTH = 20

# ヘッダ検出対象の最大範囲
MAX_HEADER_SCAN_ROWS = 6  # ヘッダ検出対象の最大行数
MAX_HEADER_SCAN_COLS = 6  # ヘッダ検出対象の最大列数

# ヘッダ判定用スコア閾値
HEADER_SCORE_THRESHOLD = 2.0

# 意味のないデータとして扱う文字列（大文字小文字区別なし）
MEANINGLESS_VALUES = ['NA', 'N/A', '#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!']

# ==================== データ構造 ====================
@dataclass
class DataElement:
    """表またはテキスト要素"""
    container_file: str  # 内包ファイル名（通常ファイルは""）
    sheet_name: str      # Excelシート名（その他は""）
    table_number: int    # 表番号（テキスト:0、表:1,2,3...）
    data_type: str       # "text" or "table"
    content: Any         # text:str, table:List[List[str]]

@dataclass
class FileResult:
    """ファイル全体の解析結果"""
    context: Dict[str, str] = field(default_factory=dict)  # {"root_file": ファイル名}
    elements: List[DataElement] = field(default_factory=list)

@dataclass
class WorksheetCache:
    """シートのキャッシュ情報"""
    sheet_name: str
    max_row: int
    max_col: int
    # 文字列化されたセル値（意味のない値は空文字に変換済み）
    cell_values_str: Dict[Tuple[int, int], str] = field(default_factory=dict)
    # セル書式情報（太字、背景色、罫線など）
    cell_formats: Dict[Tuple[int, int], Dict] = field(default_factory=dict)
    # 結合セル情報リスト (min_row, min_col, max_row, max_col)
    merged_cells: List[Tuple[int, int, int, int]] = field(default_factory=list)
    # マスターセル座標マップ (row,col) -> (master_row, master_col)
    merged_cell_map: Dict[Tuple[int, int], Tuple[int, int]] = field(default_factory=dict)
    

# ==================== ユーティリティ関数 ====================
def _is_empty_value(val: Any) -> bool:
    """
    値が空かどうか判定。空の場合True
    """
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False

def _normalize_value(val: Any) -> str:
    """
    セル値を文字列に正規化し、意味のないデータは空文字に変換
    
    Returns:
        str: 正規化された文字列
    """
    if val is None:
        return ""
    
    text = str(val).strip()
    
    # 意味のないデータの判定（大文字小文字区別なし）
    text_upper = text.upper()
    for meaningless in MEANINGLESS_VALUES:
        if text_upper == meaningless.upper():
            return ""
    
    return text


def _get_cell_format_score(cache: WorksheetCache, row_no: int, col_no: int) -> float:
    """
    セル書式からヘッダらしさのスコアを算出
    
    Returns:
        float: ヘッダスコア
    """
    fmt = cache.cell_formats.get((row_no, col_no), {})
    score = 0.0
    
    # 太字
    if fmt.get('bold', False):
        score += 1.0
    
    # 背景色
    if fmt.get('fill_type') == 'solid' and fmt.get('fill_color'):
        score += 1.0
    
    # 罫線（上下左右いずれかに太い罫線）
    borders = fmt.get('borders', {})
    if any(borders.get(side, {}).get('style') in ['medium', 'thick'] 
           for side in ['top', 'bottom', 'left', 'right']):
        score += 0.5
    
    # 結合セル（ヘッダは結合セルが多い）
    if (row_no, col_no) in cache.merged_cell_map:
        score += 1.0
    
    return score

def _get_text_pattern_score(text: str) -> float:
    """
    テキストパターンからヘッダらしさのスコアを算出
    
    Returns:
        float: ヘッダスコア
    """
    if not text:
        return 0.0
    
    score = 0.0
    
    # 短いテキスト（ヘッダは通常短い）
    if len(text) <= MAX_HEADER_LENGTH:
        score += 0.5
    
    # 数字のみは低評価
    if text.isdigit():
        score -= 1.0
    
    # 日本語・英数のラベル的な表現
    if any(c.isalpha() for c in text):
        score += 0.3
    
    return score

# ==================== キャッシュ構築 ====================
def _build_worksheet_cache(ws_data, ws_format) -> WorksheetCache:
    """
    ワークシートからWorksheetCacheを構築
    """
    cache = WorksheetCache(
        sheet_name=ws_data.title,
        max_row=min(ws_data.max_row, MAX_READ_ROWS),
        max_col=min(ws_data.max_column, MAX_READ_COLS)
    )

    # 結合セル情報（書式用ワークシートから取得）
    for merged_range in ws_format.merged_cells.ranges:
        cache.merged_cells.append((
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col
        ))

    # merged_cell_map の構築（キャッシュ時に一度だけ）
    # 例: {(3,2): (3,2), (3,3): (3,2), (4,2): (3,2), ...}
    for min_r, min_c, max_r, max_c in cache.merged_cells:
        master = (min_r, min_c)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                cache.merged_cell_map[(r, c)] = master

    # セル値と書式の読み込み
    for row in range(1, cache.max_row + 1):
        for col in range(1, cache.max_col + 1):
            # データ値の取得（data_only=Trueから）
            cell_data = ws_data.cell(row, col)
            normalized_text = _normalize_value(cell_data.value)
            cache.cell_values_str[(row, col)] = normalized_text
            
            # 書式情報の取得（data_only=Falseから）
            cell_format = ws_format.cell(row, col)
            fmt = {}
            
            # フォント情報
            if cell_format.font:
                fmt['bold'] = cell_format.font.bold or False
            
            # 塗りつぶし情報
            if cell_format.fill:
                fmt['fill_type'] = cell_format.fill.fill_type
                if hasattr(cell_format.fill, 'start_color') and cell_format.fill.start_color:
                    fmt['fill_color'] = cell_format.fill.start_color.rgb
            
            # 罫線情報
            borders = {}
            if cell_format.border:
                for side in ['top', 'bottom', 'left', 'right']:
                    border_side = getattr(cell_format.border, side, None)
                    if border_side and border_side.style:
                        borders[side] = {'style': border_side.style}
            fmt['borders'] = borders
            
            cache.cell_formats[(row, col)] = fmt
    

    return cache

# ==================== 表検出 ====================
def _detect_tables(cache: WorksheetCache) -> List[Tuple[int, int, int, int]]:
    """
    表領域を検出（min_row, min_col, max_row, max_col）
    連続する非空セル領域を探索し、最小サイズ条件を満たすものを表として認識
    
    Returns:
        List[Tuple[int, int, int, int]]: 表領域のリスト
    """
    tables = []
    visited = set()
    

    for start_row in range(1, cache.max_row + 1):
        for start_col in range(1, cache.max_col + 1):
            # すでに処理済みならスキップ
            if (start_row, start_col) in visited:
                continue
            
            # 空セルならスキップ（ただし結合セルは除く）
            if _is_empty_value(cache.cell_values_str.get((start_row, start_col))):
                # 結合セルの場合は処理を継続
                if (start_row, start_col) not in cache.merged_cell_map:
                    continue

            # 連続する非空セル領域を探索
            table_region = _expand_table_region(cache, start_row, start_col, visited)
            
            if table_region:
                min_r, min_c, max_r, max_c = table_region
                rows = max_r - min_r + 1
                cols = max_c - min_c + 1
                
                # 最小サイズチェック（ヘッダ含め3行2列以上）
                if rows >= MIN_TABLE_ROWS + 1 and cols >= MIN_TABLE_COLS:
                    tables.append(table_region)
                    # 領域を訪問済みにマーク
                    for r in range(min_r, max_r + 1):
                        for c in range(min_c, max_c + 1):
                            visited.add((r, c))
    
    return tables


def _expand_table_region(cache: WorksheetCache, start_row: int, start_col: int, 
                         visited: set) -> Optional[Tuple[int, int, int, int]]:
    """
    表領域を拡張（右方向・下方向に探索）
    
    Args:
        cache: ワークシートキャッシュ
        start_row: 開始行
        start_col: 開始列
        visited: 訪問済みセル集合
    
    Returns:
        Optional[Tuple[int, int, int, int]]: 表領域 or None
    """
    # 上方向に結合セルがあるか確認（多段ヘッダ対応）
    actual_start_row = start_row
    for check_row in range(start_row - 1, max(0, start_row - 3), -1):  # 最大3行上まで確認
        # この行に結合セルがあり、かつ下の行と列が重なっている場合は表の開始行を上げる
        has_merged = False
        for col in range(start_col, start_col + 10):  # 開始列から右10列程度チェック
            if (check_row, col) in cache.merged_cell_map:
                has_merged = True
                break
        
        if has_merged:
            actual_start_row = check_row
        else:
            break
    
    max_r, max_c = actual_start_row, start_col
    

    # まず右方向に拡張（actual_start_rowから、結合セルを考慮）
    col = start_col
    
    if (actual_start_row, start_col) in cache.merged_cell_map:
        # 結合セルの範囲を探す
        for merged_range in cache.merged_cells:
            min_mr, min_mc, max_mr, max_mc = merged_range
            if min_mr <= actual_start_row <= max_mr and min_mc <= start_col <= max_mc:
                max_c = max_mc
                col = max_mc + 1
                break
        else:
            max_c = start_col
            col = start_col + 1
    else:
        max_c = start_col
        col = start_col + 1
    
    # 結合セルの右端から継続して拡張
    while col <= cache.max_col:
        # start_rowまたはactual_start_rowのいずれかに値があれば継続
        has_value = (not _is_empty_value(cache.cell_values_str.get((start_row, col))) or
                     not _is_empty_value(cache.cell_values_str.get((actual_start_row, col))) or
                     (actual_start_row, col) in cache.merged_cell_map)
        
        if not has_value:
            break
        max_c = col
        col += 1

    
    # 下方向に拡張（各行で列数が一致するかチェック）
    row = start_row
    expected_cols = max_c - start_col + 1
    
    while row <= cache.max_row:
        # この行の非空セル数をチェック
        non_empty = 0
        for c in range(start_col, max_c + 1):
            if not _is_empty_value(cache.cell_values_str.get((row, c))):
                non_empty += 1
        
        # 空白率が高い行（50%以上空）で終了
        if non_empty < expected_cols * 0.5:
            break
        
        max_r = row
        row += 1

    return (start_row, start_col, max_r, max_c)


# ==================== ヘッダ検出 ====================
def _detect_headers(cache: WorksheetCache, table_region: Tuple[int, int, int, int]) -> Tuple[int, int]:
    """
    ヘッダ行数・列数を検出（行方向、列方向の両方）
    書式スコアとテキストパターンスコアを統合して判定
    
    Returns:
        Tuple[int, int]: (ヘッダ行数, ヘッダ列数)
    """
    min_r, min_c, max_r, max_c = table_region
    
    # 行方向ヘッダ検出（上からスキャン）
    header_rows = 0
    scan_rows = min(max_r - min_r + 1, MAX_HEADER_SCAN_ROWS)
    last_header_row = 0  # 最後にヘッダと認識された行
    
    for offset in range(scan_rows):
        row = min_r + offset
        row_score = 0.0
        cell_count = 0
        
        for col in range(min_c, max_c + 1):
            text = cache.cell_values_str.get((row, col), "")
            
            if not _is_empty_value(text):
                row_score += _get_cell_format_score(cache, row, col)
                row_score += _get_text_pattern_score(text)
                cell_count += 1
        
        # 平均スコアを計算
        avg_score = row_score / cell_count if cell_count > 0 else 0
        
        # 閾値を超えたらヘッダとして認識
        if avg_score >= HEADER_SCORE_THRESHOLD:
            header_rows = offset + 1
            last_header_row = row
        else:
            # 前の行がヘッダで、この行に結合セルがある場合は継続
            if header_rows > 0 and any((row, col) in cache.merged_cell_map for col in range(min_c, max_c + 1)):
                header_rows = offset + 1
                last_header_row = row
            else:
                break


    
    # 列方向ヘッダ検出（左からスキャン）
    header_cols = 0
    scan_cols = min(max_c - min_c + 1, MAX_HEADER_SCAN_COLS)
    
    for offset in range(scan_cols):
        col = min_c + offset
        col_score = 0.0
        cell_count = 0
        
        for row in range(min_r, max_r + 1):
            text = cache.cell_values_str.get((row, col), "")
            
            if not _is_empty_value(text):
                col_score += _get_cell_format_score(cache, row, col)
                col_score += _get_text_pattern_score(text)
                cell_count += 1
        
        avg_score = col_score / cell_count if cell_count > 0 else 0
        
        if avg_score >= HEADER_SCORE_THRESHOLD:
            header_cols = offset + 1
        else:
            break

    # 最低1行はヘッダとする
    header_rows = max(1, header_rows)
    
    return (header_rows, header_cols)

# ==================== 表データ取得 ====================
# def get_table_data(cache: WorksheetCache, table_region: Tuple[int, int, int, int]) -> Optional[List[List[str]]]:
#     """
#     表データを取得（ヘッダ結合・除外処理含む）
    
#     Args:
#         cache: ワークシートキャッシュ
#         table_region: 表領域
    
#     Returns:
#         Optional[List[List[str]]]: 表データ（ヘッダ行＋データ行）or None
#     """
#     min_r, min_c, max_r, max_c = table_region
    
#     # ヘッダ検出
#     header_rows, header_cols = _detect_headers(cache, table_region)
    
#     # データ開始位置
#     data_start_col = min_c + header_cols
#     data_start_row = min_r + header_rows
    

#     # データ行数チェック
#     actual_data_rows = max_r - data_start_row + 1
#     if actual_data_rows < MIN_TABLE_ROWS:
#         return None
    
#     # 列ヘッダの構築（多段対応）
#     headers = []
#     for col in range(data_start_col, max_c + 1):
#         header_parts = []
        
#         # ヘッダ行を上から順に走査
#         for row in range(min_r, data_start_row):
#             # 結合セルの場合は、キャッシュのマスターセルを取得
#             master_row, master_col = cache.merged_cell_map.get((row, col), (row, col))
            
#             text = cache.cell_values_str.get((master_row, master_col), "")
            
#             # 空でない場合、かつ前の行で同じ値がない場合に追加
#             if text:
#                 if not header_parts or header_parts[-1] != text:
#                     header_parts.append(text)
        
#         # 多段ヘッダを「_」で連結
#         header = "_".join(header_parts) if header_parts else f"Col{col}"
        
        
#         # 最大長制限
#         if len(header) > MAX_HEADER_LENGTH:
#             header = header[:MAX_HEADER_LENGTH]
        
#         headers.append(header)

    
#     # 除外ヘッダのチェック（部分一致）
#     exclude_indices = set()
#     for i, header in enumerate(headers):
#         for exclude in EXCLUDE_HEADERS:
#             if exclude in header:
#                 exclude_indices.add(i)
#                 break
    
#     # 除外後のヘッダ
#     filtered_headers = [h for i, h in enumerate(headers) if i not in exclude_indices]
    
#     if not filtered_headers:
#         return None
    
#     # データ行の取得
#     result = [filtered_headers]
    
#     for row in range(data_start_row, max_r + 1):
#         row_data = []
        
#         for col_idx, col in enumerate(range(data_start_col, max_c + 1)):
#             # 除外列はスキップ
#             if col_idx in exclude_indices:
#                 continue
            
#             # 結合セルの場合は、キャッシュの merged_cell_map からマスターセルを取得
#             master_row, master_col = cache.merged_cell_map.get((row, col), (row, col))
            
            
#             val = cache.cell_values_str.get((master_row, master_col), "")
#             row_data.append(val)
        
#         # 空行はスキップ
#         if any(v for v in row_data):
#             result.append(row_data)
    
#     return result if len(result) > 1 else None


def get_table_data(cache: WorksheetCache, table_region: Tuple[int, int, int, int]) -> Optional[List[List[str]]]:
    """
    表データを取得（ヘッダ結合・除外処理含む）
    
    Args:
        cache: ワークシートキャッシュ
        table_region: 表領域
    
    Returns:
        Optional[List[List[str]]]: 表データ（ヘッダ行＋データ行）or None
    """
    min_r, min_c, max_r, max_c = table_region
    
    # ヘッダ検出
    header_rows, header_cols = _detect_headers(cache, table_region)
    
    # データ開始位置
    data_start_col = min_c + header_cols
    data_start_row = min_r + header_rows
    
    # データ行数・列数チェック
    actual_data_rows = max_r - data_start_row + 1
    actual_data_cols = max_c - data_start_col + 1
    
    if actual_data_rows < MIN_TABLE_ROWS:
        return None
    
    if actual_data_cols < MIN_TABLE_COLS:
        return None
    
    # === 列ヘッダの構築と除外チェック（列方向） ===
    col_headers = []
    col_exclude_indices = set()
    
    for col in range(data_start_col, max_c + 1):
        header_parts = []
        
        # ヘッダ行を上から順に走査
        for row in range(min_r, data_start_row):
            master_row, master_col = cache.merged_cell_map.get((row, col), (row, col))
            text = cache.cell_values_str.get((master_row, master_col), "")
            
            if text:
                if not header_parts or header_parts[-1] != text:
                    header_parts.append(text)
        
        # 多段ヘッダを「_」で連結
        header = "_".join(header_parts) if header_parts else f"Col{col}"
        
        # 最大長制限
        if len(header) > MAX_HEADER_LENGTH:
            header = header[:MAX_HEADER_LENGTH]
        
        col_headers.append(header)
        
        # 除外ヘッダのチェック（部分一致）
        col_idx = col - data_start_col
        for exclude in EXCLUDE_HEADERS:
            if exclude in header:
                col_exclude_indices.add(col_idx)
                break
    
    # 除外後の列ヘッダ
    filtered_col_headers = [h for i, h in enumerate(col_headers) if i not in col_exclude_indices]
    
    if not filtered_col_headers:
        return None
    
    # === 行ヘッダの構築と除外チェック（行方向） ===
    row_exclude_indices = set()
    
    if header_cols > 0:
        for row in range(data_start_row, max_r + 1):
            row_header_parts = []
            
            # ヘッダ列を左から順に走査
            for col in range(min_c, data_start_col):
                master_row, master_col = cache.merged_cell_map.get((row, col), (row, col))
                text = cache.cell_values_str.get((master_row, master_col), "")
                
                if text:
                    if not row_header_parts or row_header_parts[-1] != text:
                        row_header_parts.append(text)
            
            # 多段ヘッダを「_」で連結
            row_header = "_".join(row_header_parts) if row_header_parts else ""
            
            # 除外ヘッダのチェック（部分一致）
            row_idx = row - data_start_row
            for exclude in EXCLUDE_HEADERS:
                if exclude in row_header:
                    row_exclude_indices.add(row_idx)
                    break
    
    # === データ行の取得 ===
    result = [filtered_col_headers]
    
    for row in range(data_start_row, max_r + 1):
        row_idx = row - data_start_row
        
        # 除外行はスキップ
        if row_idx in row_exclude_indices:
            continue
        
        row_data = []
        
        for col_idx, col in enumerate(range(data_start_col, max_c + 1)):
            # 除外列はスキップ
            if col_idx in col_exclude_indices:
                continue
            
            master_row, master_col = cache.merged_cell_map.get((row, col), (row, col))
            val = cache.cell_values_str.get((master_row, master_col), "")
            row_data.append(val)
        
        # 空行はスキップ
        if any(v for v in row_data):
            result.append(row_data)
    
    return result if len(result) > 1 else None
# ==================== テキスト抽出 ====================
def _extract_text_elements(cache: WorksheetCache, table_regions: List[Tuple[int, int, int, int]]) -> str:
    """
    表以外のテキストを抽出し、改行で連結
    
    Args:
        cache: ワークシートキャッシュ
        table_regions: 表領域のリスト
    
    Returns:
        str: 連結されたテキスト（改行区切り）
    """
    # 表領域のセット化
    table_cells = set()
    for min_r, min_c, max_r, max_c in table_regions:
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                table_cells.add((r, c))
    
    # 表外のテキストを収集
    texts = []
    
    for row in range(1, cache.max_row + 1):
        for col in range(1, cache.max_col + 1):
            # 表内のセルはスキップ
            if (row, col) in table_cells:
                continue
            
            text = cache.cell_values_str.get((row, col), "")
            if text:
                texts.append(text)
    
    # 改行で連結
    return "\n".join(texts)

# ==================== フォールバック処理 ====================
def _fallback_extract_all_data(cache: WorksheetCache) -> str:
    """
    フォールバック処理：表もテキストも取得できなかった場合、
    シート全体のデータを改行区切りで返す
    
    Args:
        cache: ワークシートキャッシュ
    
    Returns:
        str: シート全体のテキストデータ
    """
    all_texts = []
    
    for row in range(1, cache.max_row + 1):
        for col in range(1, cache.max_col + 1):
            text = cache.cell_values_str.get((row, col), "")
            if text:
                all_texts.append(text)
    
    return "\n".join(all_texts)

# ==================== メイン処理 ====================
def read_excel2(file_path: str) -> FileResult:
    """
    Excelファイルを解析してFileResultを返す
    """
    
    # FileResult初期化
    result = FileResult()
    result.context = {"root_file": os.path.basename(file_path)}
    
    # シートごとの表番号カウンター
    sheet_table_counters = {}
    
    try:
        # Excelファイルを2回読み込み
        # 1回目：データ読み込み用（data_only=True：数式の計算結果を取得）
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        
        # 2回目：書式読み込み用（data_only=False：結合セル・書式情報を取得）
        wb_format = openpyxl.load_workbook(file_path, data_only=False)
        
        for sheet_name in wb_data.sheetnames:
            # 除外シートチェック（部分一致）
            should_exclude = False
            for exclude in EXCLUDE_SHEETS:
                if exclude in sheet_name:
                    should_exclude = True
                    break
            
            if should_exclude:
                continue
            
            # このシートの表番号を初期化
            sheet_table_counters[sheet_name] = 0
            
            try:
                ws_data = wb_data[sheet_name]
                ws_format = wb_format[sheet_name]
                
                # キャッシュ構築
                cache = _build_worksheet_cache(ws_data, ws_format)
                
                # 表検出
                table_regions = _detect_tables(cache)
                
                # 表データの取得
                for table_region in table_regions:
                    table_data = get_table_data(cache, table_region)
                    
                    if table_data:
                        # 表番号をインクリメント
                        sheet_table_counters[sheet_name] += 1
                        
                        elem = DataElement(
                            container_file="",
                            sheet_name=sheet_name,
                            table_number=sheet_table_counters[sheet_name],
                            data_type="table",
                            content=table_data
                        )
                        result.elements.append(elem)
                
                # テキスト抽出
                text_content = _extract_text_elements(cache, table_regions)
                
                if text_content:
                    elem = DataElement(
                        container_file="",
                        sheet_name=sheet_name,
                        table_number=0,
                        data_type="text",
                        content=text_content
                    )
                    result.elements.append(elem)
                
                # フォールバック処理：表もテキストも取れなかった場合
                if not table_regions and not text_content:
                    fallback_data = _fallback_extract_all_data(cache)
                    
                    if fallback_data:
                        elem = DataElement(
                            container_file="",
                            sheet_name=sheet_name,
                            table_number=0,
                            data_type="text",
                            content=fallback_data
                        )
                        result.elements.append(elem)
            
            except Exception as e:
                print(e)
                print(f"ファイル '{file_path}' のFilrresult生成エラー: {e}")
                continue
        
        wb_data.close()
        wb_format.close()
    
    except Exception as e:
        print(f"ファイル '{file_path}' の読み込みエラー: {e}")
    
    return result


# ==================== 使用例 ====================
if __name__ == "__main__":
    # 使用例
    file_path = "test.xlsx"
    
    result = read_excel2(file_path)
    
    print(f"ファイル: {result.context['root_file']}")
    print(f"要素数: {len(result.elements)}\n")
    
    for i, elem in enumerate(result.elements, 1):
        print(f"--- 要素 {i} ---")
        print(f"container_file: {elem.container_file}")
        print(f"sheet_name: {elem.sheet_name}")
        print(f"table_number: {elem.table_number}")
        print(f"data_type: {elem.data_type}")
        
        if elem.data_type == "table":
            print(f"行数: {len(elem.content)}")
            print(f"列数: {len(elem.content[0]) if elem.content else 0}")
            print("ヘッダ:", elem.content[0] if elem.content else [])
            print("データ例（最大3行）:")
            for row in elem.content[1:4]:
                print("  ", row)
        else:
            preview = elem.content[:100] + "..." if len(elem.content) > 100 else elem.content
            print(f"テキスト: {preview}")
        
        print()